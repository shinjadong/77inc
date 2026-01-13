"""
Excel 내보내기 API
Supabase Storage 연동
"""
import logging
from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.services.excel_export import ExcelExportService
from app.services.supabase_storage import get_storage_service

router = APIRouter()
logger = logging.getLogger(__name__)


@router.get("/months")
async def get_available_months(db: Session = Depends(get_db)):
    """거래가 있는 월 목록 조회"""
    service = ExcelExportService(db)
    months = service.get_available_months()

    return {"months": months}


@router.get("/monthly/{year}/{month}")
async def export_monthly(
    year: int,
    month: int,
    save_to_storage: bool = Query(False, description="Supabase Storage에 저장"),
    db: Session = Depends(get_db),
):
    """월별 Excel 내보내기"""
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="잘못된 월입니다 (1-12)")

    service = ExcelExportService(db)
    excel_bytes = service.export_monthly(year, month)

    filename = f"칠칠기업_법인카드_{year}_{month:02d}.xlsx"

    # Storage에 저장 (옵션)
    storage_info = None
    if save_to_storage:
        try:
            storage = get_storage_service()
            storage_info = storage.upload_excel_export(excel_bytes, filename)
            logger.info(f"Exported file saved to storage: {storage_info['path']}")
        except Exception as e:
            logger.error(f"Failed to save to storage: {e}")

    encoded_filename = quote(filename)

    response = StreamingResponse(
        BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )

    # Storage URL을 헤더에 포함 (저장된 경우)
    if storage_info:
        response.headers["X-Storage-URL"] = storage_info["url"]

    return response


@router.get("/all")
async def export_all(
    save_to_storage: bool = Query(False, description="Supabase Storage에 저장"),
    db: Session = Depends(get_db),
):
    """전체 Excel 내보내기"""
    service = ExcelExportService(db)
    excel_bytes = service.export_all()

    filename = "칠칠기업_법인카드.xlsx"

    # Storage에 저장 (옵션)
    storage_info = None
    if save_to_storage:
        try:
            storage = get_storage_service()
            storage_info = storage.upload_excel_export(excel_bytes, filename)
            logger.info(f"Exported file saved to storage: {storage_info['path']}")
        except Exception as e:
            logger.error(f"Failed to save to storage: {e}")

    encoded_filename = quote(filename)

    response = StreamingResponse(
        BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )

    if storage_info:
        response.headers["X-Storage-URL"] = storage_info["url"]

    return response


@router.get("/files")
async def list_export_files():
    """저장된 내보내기 파일 목록"""
    try:
        storage = get_storage_service()
        files = storage.list_files("exports")
        return {"files": files}
    except Exception as e:
        logger.error(f"Failed to list files: {e}")
        raise HTTPException(status_code=500, detail="파일 목록 조회 실패")


@router.get("/card/{card_number}")
async def export_by_card(
    card_number: str,
    year: int = None,
    month: int = None,
    save_to_storage: bool = Query(False, description="Supabase Storage에 저장"),
    db: Session = Depends(get_db),
):
    """
    카드별 개별 Excel 내보내기

    - card_number: 카드번호 뒷4자리 (예: 3987)
    - year/month: 특정 월 필터링 (선택)
    """
    from app.repositories.card_repo import CardRepository
    from app.models.transaction import Transaction
    from app.models.card import Card
    from sqlalchemy import extract

    # 카드 확인
    card_repo = CardRepository(db)
    card = card_repo.get_by_number(card_number)
    if not card:
        raise HTTPException(status_code=404, detail=f"카드번호 {card_number}를 찾을 수 없습니다")

    # 거래 조회
    query = db.query(Transaction).filter(Transaction.card_id == card.id)

    if year and month:
        query = query.filter(
            extract('year', Transaction.transaction_date) == year,
            extract('month', Transaction.transaction_date) == month
        )

    transactions = query.order_by(Transaction.transaction_date).all()

    # 단일 카드용 Excel 생성
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = card_number

    # 헤더
    headers = ["결제일자", "가맹점명", "이용금액", "사용용도"]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    column_widths = [12, 30, 15, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # 데이터
    for row_idx, tx in enumerate(transactions, 2):
        ws.cell(row=row_idx, column=1, value=tx.transaction_date).number_format = 'YYYY-MM-DD'
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=1).border = thin_border

        ws.cell(row=row_idx, column=2, value=tx.merchant_name).border = thin_border

        ws.cell(row=row_idx, column=3, value=tx.amount).number_format = '#,##0'
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right")
        ws.cell(row=row_idx, column=3).border = thin_border

        ws.cell(row=row_idx, column=4, value=tx.usage_description or "").border = thin_border

    # BytesIO로 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    excel_bytes = output.read()

    # 파일명 생성
    if year and month:
        filename = f"카드_{card_number}_{year}_{month:02d}.xlsx"
    else:
        filename = f"카드_{card_number}_전체.xlsx"

    # Storage 저장 (옵션)
    storage_info = None
    if save_to_storage:
        try:
            storage = get_storage_service()
            storage_info = storage.upload_excel_export(excel_bytes, filename)
        except Exception as e:
            logger.error(f"Failed to save to storage: {e}")

    encoded_filename = quote(filename)

    response = StreamingResponse(
        BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )

    if storage_info:
        response.headers["X-Storage-URL"] = storage_info["url"]

    return response


@router.get("/card/{card_number}/stats")
async def get_card_stats(
    card_number: str,
    db: Session = Depends(get_db),
):
    """카드별 거래 통계"""
    from app.repositories.card_repo import CardRepository
    from app.models.transaction import Transaction
    from sqlalchemy import func, extract

    card_repo = CardRepository(db)
    card = card_repo.get_by_number(card_number)
    if not card:
        raise HTTPException(status_code=404, detail=f"카드번호 {card_number}를 찾을 수 없습니다")

    # 전체 통계
    total = db.query(func.count(Transaction.id)).filter(
        Transaction.card_id == card.id
    ).scalar()

    matched = db.query(func.count(Transaction.id)).filter(
        Transaction.card_id == card.id,
        Transaction.usage_description.isnot(None),
        Transaction.usage_description != ""
    ).scalar()

    pending = total - matched

    # 월별 통계
    monthly = db.query(
        extract('year', Transaction.transaction_date).label('year'),
        extract('month', Transaction.transaction_date).label('month'),
        func.count(Transaction.id).label('count'),
        func.sum(Transaction.amount).label('total_amount')
    ).filter(
        Transaction.card_id == card.id
    ).group_by(
        extract('year', Transaction.transaction_date),
        extract('month', Transaction.transaction_date)
    ).order_by(
        extract('year', Transaction.transaction_date).desc(),
        extract('month', Transaction.transaction_date).desc()
    ).all()

    return {
        "card_number": card_number,
        "card_name": card.card_name,
        "total_transactions": total,
        "matched": matched,
        "pending": pending,
        "match_rate": round(matched / total * 100, 1) if total > 0 else 0,
        "monthly": [
            {
                "year": int(m.year),
                "month": int(m.month),
                "count": m.count,
                "total_amount": m.total_amount or 0
            }
            for m in monthly
        ]
    }
