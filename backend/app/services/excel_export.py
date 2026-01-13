"""
Excel 내보내기 서비스
세무사 제출용 법인카드 내역 Excel 파일 생성
"""
from io import BytesIO
from datetime import date
from typing import Optional, Dict, List
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import extract

from app.models.transaction import Transaction
from app.models.card import Card


class ExcelExportService:
    """Excel 내보내기 서비스"""

    # 카드 시트 순서 (고정)
    CARD_ORDER = ["3987", "4985", "6902", "6974", "9980", "6911"]

    # 컬럼 헤더
    HEADERS = ["결제일자", "가맹점명", "이용금액", "사용용도"]

    def __init__(self, db: Session):
        self.db = db

    def export_monthly(self, year: int, month: int) -> bytes:
        """월별 데이터 내보내기"""
        # 해당 월의 거래 조회
        transactions = self.db.query(Transaction).filter(
            extract('year', Transaction.transaction_date) == year,
            extract('month', Transaction.transaction_date) == month
        ).order_by(
            Transaction.card_id,
            Transaction.transaction_date
        ).all()

        # 카드별로 그룹화
        transactions_by_card = self._group_by_card(transactions)

        return self._create_workbook(transactions_by_card)

    def export_all(self) -> bytes:
        """전체 데이터 내보내기"""
        # 모든 거래 조회
        transactions = self.db.query(Transaction).order_by(
            Transaction.card_id,
            Transaction.transaction_date
        ).all()

        # 카드별로 그룹화
        transactions_by_card = self._group_by_card(transactions)

        return self._create_workbook(transactions_by_card)

    def export_date_range(
        self,
        start_date: date,
        end_date: date
    ) -> bytes:
        """기간별 데이터 내보내기"""
        transactions = self.db.query(Transaction).filter(
            Transaction.transaction_date >= start_date,
            Transaction.transaction_date <= end_date
        ).order_by(
            Transaction.card_id,
            Transaction.transaction_date
        ).all()

        transactions_by_card = self._group_by_card(transactions)

        return self._create_workbook(transactions_by_card)

    def _group_by_card(
        self,
        transactions: List[Transaction]
    ) -> Dict[str, List[Transaction]]:
        """카드별로 거래 그룹화"""
        # 카드 ID -> 카드번호 매핑 조회
        cards = self.db.query(Card).filter(Card.is_active == True).all()
        card_number_map = {c.id: c.card_number for c in cards}

        # 카드번호별로 그룹화
        grouped: Dict[str, List[Transaction]] = defaultdict(list)
        for tx in transactions:
            card_number = card_number_map.get(tx.card_id, "unknown")
            grouped[card_number].append(tx)

        return dict(grouped)

    def _create_workbook(
        self,
        transactions_by_card: Dict[str, List[Transaction]]
    ) -> bytes:
        """워크북 생성"""
        wb = Workbook()

        # 기본 시트 제거
        if wb.active:
            wb.remove(wb.active)

        # 카드 순서대로 시트 생성
        for card_number in self.CARD_ORDER:
            ws = wb.create_sheet(title=card_number)
            transactions = transactions_by_card.get(card_number, [])
            self._fill_sheet(ws, transactions)

        # BytesIO로 저장
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output.read()

    def _fill_sheet(self, ws, transactions: List[Transaction]):
        """시트에 데이터 채우기"""
        # 스타일 정의
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 컬럼 너비 설정
        column_widths = [12, 30, 15, 40]  # 결제일자, 가맹점명, 이용금액, 사용용도
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 헤더 작성
        for col, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 데이터 작성
        for row_idx, tx in enumerate(transactions, 2):
            # 결제일자
            date_cell = ws.cell(row=row_idx, column=1, value=tx.transaction_date)
            date_cell.number_format = 'YYYY-MM-DD'
            date_cell.alignment = Alignment(horizontal="center")
            date_cell.border = thin_border

            # 가맹점명
            merchant_cell = ws.cell(row=row_idx, column=2, value=tx.merchant_name)
            merchant_cell.border = thin_border

            # 이용금액
            amount_cell = ws.cell(row=row_idx, column=3, value=tx.amount)
            amount_cell.number_format = '#,##0'
            amount_cell.alignment = Alignment(horizontal="right")
            amount_cell.border = thin_border

            # 사용용도
            usage_cell = ws.cell(row=row_idx, column=4, value=tx.usage_description or "")
            usage_cell.border = thin_border

        # 데이터가 없는 경우 빈 행 표시
        if not transactions:
            ws.cell(row=2, column=1, value="데이터 없음")
            ws.merge_cells('A2:D2')
            ws['A2'].alignment = Alignment(horizontal="center")

    def get_available_months(self) -> List[Dict]:
        """거래가 있는 월 목록 조회"""
        from sqlalchemy import func, distinct

        result = self.db.query(
            extract('year', Transaction.transaction_date).label('year'),
            extract('month', Transaction.transaction_date).label('month'),
            func.count(Transaction.id).label('count')
        ).group_by(
            extract('year', Transaction.transaction_date),
            extract('month', Transaction.transaction_date)
        ).order_by(
            extract('year', Transaction.transaction_date).desc(),
            extract('month', Transaction.transaction_date).desc()
        ).all()

        return [
            {
                "year": int(r.year),
                "month": int(r.month),
                "count": r.count
            }
            for r in result
        ]
