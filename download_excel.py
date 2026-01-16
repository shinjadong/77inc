#!/usr/bin/env python3
"""
Supabase DB에서 거래 데이터를 가져와 엑셀 파일로 내보내기
"""
import os
import sys
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from supabase import create_client, Client

# Supabase 설정
SUPABASE_URL = "https://kxcvsgecefbzoiczyxsp.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Imt4Y3ZzZ2VjZWZiem9pY3p5eHNwIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NjgyMzcwNDgsImV4cCI6MjA4MzgxMzA0OH0.LpzRg_uzhauq-eyp1iNEVyM37wZxU2LmOUt6OAgwUBI"

# 카드 순서
CARD_ORDER = ['3987', '4985', '6902', '6974', '9980', '6911', '0981', '9904']

def fetch_transactions():
    """Supabase에서 거래 데이터 조회"""
    print("Supabase에서 데이터 조회 중...")
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

    # transactions + cards 조인
    response = supabase.table('transactions').select(
        'id, transaction_date, merchant_name, amount, usage_description, '
        'additional_notes, tax_category, cards(card_number, card_name)'
    ).order('transaction_date').execute()

    if not response.data:
        print("조회된 데이터가 없습니다.")
        return []

    # 데이터 가공
    transactions = []
    for row in response.data:
        card_info = row.get('cards', {})
        transactions.append({
            'card_number': card_info.get('card_number', ''),
            'card_name': card_info.get('card_name', ''),
            'transaction_date': row['transaction_date'],
            'merchant_name': row['merchant_name'],
            'amount': row['amount'],
            'usage_description': row.get('usage_description', ''),
            'additional_notes': row.get('additional_notes', ''),
            'tax_category': row.get('tax_category', '')
        })

    print(f"총 {len(transactions)}건의 거래 조회 완료")
    return transactions

def create_excel(transactions, output_path):
    """엑셀 파일 생성"""
    print(f"\n엑셀 파일 생성 중: {output_path}")

    # DataFrame으로 변환
    df = pd.DataFrame(transactions)

    # 카드별로 그룹화
    grouped = {}
    for card_number in CARD_ORDER:
        grouped[card_number] = df[df['card_number'] == card_number].copy()

    # ExcelWriter 생성
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for card_number in CARD_ORDER:
            card_df = grouped[card_number]

            if len(card_df) == 0:
                print(f"  - 시트 {card_number}: 데이터 없음 (건너뜀)")
                continue

            # 날짜순 정렬
            card_df = card_df.sort_values('transaction_date')

            # 필요한 컬럼만 선택
            export_df = card_df[[
                'transaction_date', 'merchant_name', 'amount',
                'usage_description', 'additional_notes', 'tax_category'
            ]].copy()

            # 컬럼명 변경
            export_df.columns = ['결제일자', '가맹점명', '이용금액', '사용용도', '추가메모', '세금분류']

            # 시트에 쓰기
            sheet_name = card_number
            export_df.to_excel(writer, sheet_name=sheet_name, index=False)

            # 워크시트 가져오기
            ws = writer.sheets[sheet_name]

            # 헤더 스타일
            header_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            header_font = Font(bold=True, size=11)
            header_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 헤더 행 스타일 적용
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

            # 데이터 행 스타일 및 포맷
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                # 날짜 포맷 (A열)
                row[0].number_format = 'YYYY-MM-DD'
                row[0].alignment = Alignment(horizontal='center')

                # 금액 포맷 (C열)
                row[2].number_format = '#,##0'
                row[2].alignment = Alignment(horizontal='right')

                # 테두리
                for cell in row:
                    cell.border = border

            # 컬럼 너비 조정
            ws.column_dimensions['A'].width = 12  # 결제일자
            ws.column_dimensions['B'].width = 30  # 가맹점명
            ws.column_dimensions['C'].width = 15  # 이용금액
            ws.column_dimensions['D'].width = 30  # 사용용도
            ws.column_dimensions['E'].width = 40  # 추가메모
            ws.column_dimensions['F'].width = 20  # 세금분류

            # 헤더 행 높이
            ws.row_dimensions[1].height = 25

            print(f"  - 시트 {card_number}: {len(card_df)}건 작성 완료")

    print(f"\n✅ 엑셀 파일 생성 완료: {output_path}")

def main():
    # 출력 디렉토리
    output_dir = '/home/tlswkehd/77inc/output'
    os.makedirs(output_dir, exist_ok=True)

    # 파일명 생성 (날짜 포함)
    today = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = os.path.join(output_dir, f'칠칠기업_법인카드_{today}.xlsx')

    try:
        # 1. 데이터 조회
        transactions = fetch_transactions()

        if not transactions:
            print("데이터가 없어 엑셀 파일을 생성할 수 없습니다.")
            return

        # 2. 엑셀 생성
        create_excel(transactions, output_file)

        # 3. 요약 출력
        print("\n" + "="*60)
        print("📊 다운로드 요약")
        print("="*60)

        df = pd.DataFrame(transactions)
        for card_number in CARD_ORDER:
            card_df = df[df['card_number'] == card_number]
            if len(card_df) > 0:
                total_amount = card_df['amount'].sum()
                print(f"카드 {card_number}: {len(card_df)}건, 총액 {total_amount:,}원")

        print(f"\n✅ 총 {len(transactions)}건의 거래가 엑셀로 저장되었습니다.")
        print(f"📁 파일 위치: {output_file}")

    except Exception as e:
        print(f"\n❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
