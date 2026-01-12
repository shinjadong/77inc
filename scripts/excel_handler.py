#!/usr/bin/env python3
"""
Excel 파일 처리 모듈
칠칠기업_법인카드.xlsx 읽기/쓰기
"""
from typing import List, Tuple, Optional
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelHandler:
    """메인 Excel 파일 핸들러"""

    # 컬럼 순서
    COLUMNS = ["결제일자", "가맹점명", "이용금액", "사용용도"]

    def __init__(self, file_path: str = "/home/tlswk/77corp/칠칠기업_법인카드.xlsx"):
        self.file_path = Path(file_path)
        self.workbook = None
        self._existing_data = {}  # 시트별 기존 데이터 캐시

        if not self.file_path.exists():
            raise FileNotFoundError(f"메인 파일을 찾을 수 없습니다: {file_path}")

    def open(self):
        """워크북 열기"""
        self.workbook = load_workbook(self.file_path)
        self._cache_existing_data()

    def _cache_existing_data(self):
        """기존 데이터 캐시 (중복 검사용)"""
        for sheet_name in self.workbook.sheetnames:
            if sheet_name.isdigit():  # 카드 시트만
                df = pd.read_excel(self.file_path, sheet_name=sheet_name)
                self._existing_data[sheet_name] = df

    def is_duplicate(self, sheet_name: str, date: str, merchant: str, amount: int) -> bool:
        """
        중복 거래 검사

        Args:
            sheet_name: 카드 시트명
            date: 결제일자 (YYYY-MM-DD)
            merchant: 가맹점명
            amount: 금액

        Returns:
            중복 여부
        """
        if sheet_name not in self._existing_data:
            return False

        df = self._existing_data[sheet_name]

        for _, row in df.iterrows():
            # 결제일자 비교 (날짜만)
            row_date = str(row.get("결제일자", ""))[:10]
            if row_date != date[:10]:
                continue

            # 가맹점명 비교
            row_merchant = str(row.get("가맹점명", "")).strip()
            if row_merchant != merchant:
                continue

            # 금액 비교
            try:
                row_amount = int(row.get("이용금액", 0))
            except:
                row_amount = 0

            if row_amount == amount:
                return True

        return False

    def append_row(self, sheet_name: str, date: str, merchant: str, amount: int, usage: str):
        """
        시트에 새 행 추가

        Args:
            sheet_name: 카드 시트명
            date: 결제일자
            merchant: 가맹점명
            amount: 금액
            usage: 사용용도
        """
        if sheet_name not in self.workbook.sheetnames:
            # 시트가 없으면 생성
            ws = self.workbook.create_sheet(sheet_name)
            ws.append(self.COLUMNS)
        else:
            ws = self.workbook[sheet_name]

        # 새 행 추가
        ws.append([date, merchant, amount, usage])

        # 캐시 업데이트
        if sheet_name not in self._existing_data:
            self._existing_data[sheet_name] = pd.DataFrame(columns=self.COLUMNS)

        new_row = pd.DataFrame([[date, merchant, amount, usage]], columns=self.COLUMNS)
        self._existing_data[sheet_name] = pd.concat(
            [self._existing_data[sheet_name], new_row],
            ignore_index=True
        )

    def get_sheet_data(self, sheet_name: str) -> pd.DataFrame:
        """시트 데이터 조회"""
        if sheet_name in self._existing_data:
            return self._existing_data[sheet_name].copy()
        return pd.DataFrame(columns=self.COLUMNS)

    def get_all_cards_data(self) -> dict:
        """모든 카드 시트 데이터 조회"""
        result = {}
        for sheet_name in self.workbook.sheetnames:
            if sheet_name.isdigit():
                result[sheet_name] = self.get_sheet_data(sheet_name)
        return result

    def save(self, backup: bool = True):
        """
        파일 저장

        Args:
            backup: 백업 생성 여부
        """
        if backup:
            # 백업 생성
            backup_dir = self.file_path.parent / "backup"
            backup_dir.mkdir(exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = backup_dir / f"법인카드_backup_{timestamp}.xlsx"
            self.workbook.save(backup_path)
            print(f"백업 생성: {backup_path.name}")

        # 원본 저장
        self.workbook.save(self.file_path)
        print(f"저장 완료: {self.file_path.name}")

    def close(self):
        """워크북 닫기"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None

    def __enter__(self):
        self.open()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()


class PendingReportHandler:
    """미분류 리포트 핸들러"""

    COLUMNS = ["결제일자", "가맹점명", "이용금액", "카드", "업종", "최종_사용용도"]

    def __init__(self, output_dir: str = "/home/tlswk/77corp/미분류"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.pending_items = []

    def add(self, date: str, merchant: str, amount: int, card: str, industry: str = ""):
        """미분류 항목 추가"""
        self.pending_items.append({
            "결제일자": date,
            "가맹점명": merchant,
            "이용금액": amount,
            "카드": card,
            "업종": industry,
            "최종_사용용도": ""  # 수동 입력용 빈 컬럼
        })

    def save(self) -> Optional[str]:
        """
        미분류 리포트 저장

        Returns:
            저장된 파일 경로 (항목이 없으면 None)
        """
        if not self.pending_items:
            return None

        df = pd.DataFrame(self.pending_items, columns=self.COLUMNS)
        timestamp = datetime.now().strftime("%Y%m%d")
        output_path = self.output_dir / f"pending_{timestamp}.xlsx"

        df.to_excel(output_path, index=False)
        print(f"미분류 리포트 저장: {output_path.name} ({len(self.pending_items)}건)")

        return str(output_path)

    @property
    def count(self) -> int:
        return len(self.pending_items)


if __name__ == "__main__":
    # 테스트
    handler = ExcelHandler()
    handler.open()

    print("\n시트 목록:")
    for sheet in handler.workbook.sheetnames:
        if sheet.isdigit():
            df = handler.get_sheet_data(sheet)
            print(f"  [{sheet}] {len(df)}건")

    # 중복 검사 테스트
    print("\n중복 검사 테스트:")
    is_dup = handler.is_duplicate("3987", "2025-07-07", "맥도날드 안산고잔DT점", 8400)
    print(f"  맥도날드 안산고잔DT점 (2025-07-07, 8400): {'중복' if is_dup else '신규'}")

    handler.close()
